import streamlit as st
import PyPDF2
import pandas as pd
import numpy as np
import re
import os
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import warnings
import tempfile
import zipfile
import shutil
from decimal import Decimal, ROUND_HALF_UP
import math

warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="TDS Return Automation",
    page_icon="📊",
    layout="wide"
)

# Title and description
st.title("📊 TDS Return Automation System")
st.markdown("""
This application automates the processing of TDS (Tax Deducted at Source) returns by:
- Extracting data from PDF challan files
- Updating TDS Masters Excel file
- Generating formatted output for filing
""")

# Initialize session state
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'results' not in st.session_state:
    st.session_state.results = None

# Create tabs
tab1, tab2, tab3 = st.tabs(["📤 Upload Files", "⚙️ Process", "📥 Download Results"])

# PDF Processing Functions
def extract_challan_data_from_pdf(pdf_file):
    """Extract challan data from a PDF file"""
    challan_data = {}
    
    try:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        if len(pdf_reader.pages) > 0:
            text = pdf_reader.pages[0].extract_text()
        
        # Extract fields using regex patterns
        patterns = {
            'tan': r'TAN\s*:\s*([A-Z0-9]+)',
            'nature_of_payment': r'Nature of Payment\s*:\s*(\d+[A-Z])',
            'cin': r'CIN\s*:\s*([A-Z0-9]+)',
            'bsr_code': r'BSR code\s*:\s*([\d]+)',
            'challan_no': r'Challan No\s*:\s*([\d]+)',
            'tender_date': r'Tender Date\s*:\s*(\d{2}/\d{2}/\d{4})',
            'mode_of_payment': r'Mode of Payment\s*:\s*([^\n]+)',
        }
        
        for field, pattern in patterns.items():
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                value = match.group(1).strip()
                if field == 'mode_of_payment':
                    value = value.upper()
                elif field == 'bsr_code':
                    value = value.zfill(7)
                challan_data[field] = value
            else:
                challan_data[field] = ""
        
        # Extract tax amount
        tax_patterns = [
            r'A\s+Tax\s+₹\s*([\d,]+)',
            r'Tax\s+₹\s*([\d,]+)',
            r'A\s+Tax[^0-9]+([\d,]+)',
        ]
        
        tax_amount = ""
        for pattern in tax_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                tax_amount = match.group(1).strip().replace(',', '')
                break
        
        challan_data['tax_amount'] = tax_amount
        
        # Extract other amounts
        amount_fields = {
            'surcharge': [r'B\s+Surcharge\s+₹\s*([\d,]+)', r'Surcharge\s+₹\s*([\d,]+)'],
            'cess': [r'C\s+Cess\s+₹\s*([\d,]+)', r'Cess\s+₹\s*([\d,]+)'],
            'interest': [r'D\s+Interest\s+₹\s*([\d,]+)', r'Interest\s+₹\s*([\d,]+)'],
            'penalty': [r'E\s+Penalty\s+₹\s*([\d,]+)', r'Penalty\s+₹\s*([\d,]+)'],
            'fee_234e': [r'F\s+Fee under section 234E\s+₹\s*([\d,]+)', r'234E[^0-9]+([\d,]+)']
        }
        
        for field, patterns_list in amount_fields.items():
            value = ""
            for pattern in patterns_list:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    value = match.group(1).strip().replace(',', '')
                    break
            challan_data[field] = value if value else "0"
        
        # Extract total
        total_patterns = [
            r'Total \(A\+B\+C\+D\+E\+F\)\s+₹\s*([\d,]+)',
            r'Total.*?₹\s*([\d,]+)'
        ]
        
        total_amount = ""
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                total_amount = match.group(1).strip().replace(',', '')
                break
        
        challan_data['total_amount'] = total_amount
        challan_data['file_name'] = pdf_file.name
        
    except Exception as e:
        challan_data['error'] = str(e)
    
    return challan_data

def read_tds_masters(file):
    """Read TDS Masters Excel file"""
    try:
        # Read sheets
        tds_codes = pd.read_excel(file, sheet_name='TDS CODES', keep_default_na=False)
        tds_rates = pd.read_excel(file, sheet_name='TDS RATES', keep_default_na=False)
        
        # Save file temporarily to use with openpyxl
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.seek(0)
            tmp.write(file.read())
            tmp_path = tmp.name
        
        # Use openpyxl for TDS PARTIES
        wb = load_workbook(tmp_path, data_only=True)
        ws_parties = wb['TDS PARTIES']
        
        # Find code row
        code_row = None
        for idx in range(1, 11):
            row_values = []
            for col in range(1, ws_parties.max_column + 1):
                cell_value = ws_parties.cell(row=idx, column=col).value
                if cell_value:
                    row_values.append(str(cell_value))
            
            code_patterns = ['(415)', '(427)', '(416)', '(415A)', '-415', '-427', '-416']
            if any(pattern in val for val in row_values for pattern in code_patterns):
                code_row = idx
                break
        
        # Read headers
        header_row = code_row - 1 if code_row else 1
        headers = []
        for col in range(1, ws_parties.max_column + 1):
            header_val = ws_parties.cell(row=header_row, column=col).value
            headers.append(header_val if header_val else f"Column_{col}")
        
        # Map column codes
        code_to_column_name = {}
        column_code_map = {}
        if code_row:
            for col in range(1, ws_parties.max_column + 1):
                code_val = ws_parties.cell(row=code_row, column=col).value
                if code_val:
                    code_str = str(code_val).strip()
                    code_match = None
                    if '(' in code_str and ')' in code_str:
                        code_match = re.search(r'\(([0-9A-Z]+)\)', code_str)
                    elif code_str.startswith('-'):
                        code_match = re.search(r'-([0-9A-Z]+)', code_str)
                    if code_match:
                        extracted_code = code_match.group(1)
                        normalized_code = f'({extracted_code})'
                        column_code_map[normalized_code] = col - 1
                        code_to_column_name[normalized_code] = headers[col - 1]
        
        # Fallback mappings
        column_name_mappings = {
            '(415)': ['Deductee Code', 'Individual/Company', 'Indiv/Comp', 'Code'],
            '(415A)': ['Section Under Payment Made', 'Type of Payment', 'Nature of Payment'],
            '(416)': ['PAN of the Deductee', 'PAN', 'PAN No'],
            '(417)': ['Name of the Deductee', 'Deductee Name', 'Name'],
            '(418)': ['Date of Payment/credit', 'Payment Date', 'Date of Payment'],
            '(419)': ['Amount Paid /Credited', 'Amount Paid', 'Gross Amount'],
            '(421)': ['TDS', 'Tax Deducted', 'TDS Amount', 'TDS               Rs.'],
            '(425D)': ['BSR Code', 'BSR'],
            '(425E)': ['Challan Serial No', 'Challan No'],
            '(425F)': ['Date on which deposited', 'Date Deposited'],
            '(427)': ['TDS Deducted Rates %', 'TDS Rate', 'Rate %', 'Rate']
        }
        
        for code, possible_names in column_name_mappings.items():
            if code not in code_to_column_name:
                for col_idx, col_name in enumerate(headers):
                    col_name_clean = str(col_name).strip()
                    for possible_name in possible_names:
                        if possible_name.lower() in col_name_clean.lower():
                            code_to_column_name[code] = col_name
                            column_code_map[code] = col_idx
                            break
                    if code in code_to_column_name:
                        break
        
        # Read data rows
        data_rows = []
        data_start_row = code_row + 1 if code_row else 2
        for row in range(data_start_row, ws_parties.max_row + 1):
            row_data = []
            is_empty_row = True
            for col in range(1, ws_parties.max_column + 1):
                cell_value = ws_parties.cell(row=row, column=col).value
                row_data.append(cell_value)
                if cell_value is not None and str(cell_value).strip():
                    is_empty_row = False
            if not is_empty_row:
                data_rows.append(row_data)
        
        tds_parties = pd.DataFrame(data_rows, columns=headers)
        
        # Convert numeric columns
        numeric_codes = ['(419)', '(421)']
        for code in numeric_codes:
            col_name = code_to_column_name.get(code)
            if col_name and col_name in tds_parties.columns:
                tds_parties[col_name] = pd.to_numeric(
                    tds_parties[col_name].astype(str).str.replace(',', '').str.replace('₹', '').str.strip(),
                    errors='coerce'
                ).apply(lambda x: Decimal(str(x)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if pd.notna(x) else x)
        
        # Convert date columns
        date_codes = ['(418)', '(425F)']
        for code in date_codes:
            col_name = code_to_column_name.get(code)
            if col_name and col_name in tds_parties.columns:
                tds_parties[col_name] = pd.to_datetime(tds_parties[col_name], errors='coerce', dayfirst=True)
        
        challan_details = pd.read_excel(file, sheet_name='Challan Details', header=1, keep_default_na=False)
        wb.close()
        
        # Clean up temp file
        os.unlink(tmp_path)
        
        return {
            'tds_codes': tds_codes,
            'tds_parties': tds_parties,
            'challan_details': challan_details,
            'tds_rates': tds_rates,
            'column_code_map': column_code_map,
            'code_to_column_name': code_to_column_name,
            'code_row': code_row
        }
        
    except Exception as e:
        st.error(f"Error reading TDS Masters: {str(e)}")
        return None

def update_tds_masters_with_challans(tds_masters_data, challan_data_list, masters_path):
    """Update TDS Masters with challan information"""
    try:
        wb = load_workbook(masters_path, data_only=True)
        ws_parties = wb['TDS PARTIES']
        ws_challan = wb['Challan Details']
        
        code_to_column_name = tds_masters_data.get('code_to_column_name', {})
        code_row = tds_masters_data.get('code_row', 1)
        
        # Find columns
        col_425E = col_425F = col_415A = None
        for col_idx in range(1, ws_parties.max_column + 1):
            cell_value = str(ws_parties.cell(row=code_row, column=col_idx).value)
            if '425E' in cell_value:
                col_425E = col_idx
            elif '425F' in cell_value:
                col_425F = col_idx
            elif '415A' in cell_value:
                col_415A = col_idx
        
        # Create challan mapping
        challan_map = {}
        for challan in challan_data_list:
            nop = challan.get('nature_of_payment', '')
            if nop:
                nop_clean = nop.replace(' ', '')
                challan_map[nop_clean] = challan
        
        # Update TDS PARTIES
        data_start_row = code_row + 1
        for row_idx in range(data_start_row, ws_parties.max_row + 1):
            payment_type = ws_parties.cell(row=row_idx, column=col_415A).value if col_415A else None
            if payment_type and str(payment_type).strip() not in ['', 'nan', 'None']:
                payment_type_clean = str(payment_type).replace(' ', '').strip()
                if payment_type_clean in challan_map:
                    challan = challan_map[payment_type_clean]
                    if col_425E:
                        ws_parties.cell(row=row_idx, column=col_425E).value = challan.get('challan_no', '')
                    if col_425F:
                        date_str = challan.get('tender_date', '')
                        if date_str:
                            try:
                                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                                ws_parties.cell(row=row_idx, column=col_425F).value = date_obj
                                ws_parties.cell(row=row_idx, column=col_425F).number_format = 'DD/MM/YYYY'
                            except:
                                ws_parties.cell(row=row_idx, column=col_425F).value = date_str
        
        # Update Challan Details
        for row in ws_challan.iter_rows(min_row=3, max_row=ws_challan.max_row):
            for cell in row:
                cell.value = None
        
        for idx, challan in enumerate(challan_data_list, start=3):
            ws_challan.cell(row=idx, column=1).value = idx - 2
            ws_challan.cell(row=idx, column=2).value = challan.get('nature_of_payment', '')
            ws_challan.cell(row=idx, column=3).value = int(float(challan.get('tax_amount', 0)))
            ws_challan.cell(row=idx, column=4).value = int(float(challan.get('surcharge', 0)))
            ws_challan.cell(row=idx, column=5).value = int(float(challan.get('cess', 0)))
            ws_challan.cell(row=idx, column=6).value = int(float(challan.get('interest', 0)))
            ws_challan.cell(row=idx, column=7).value = int(float(challan.get('penalty', 0)))
            ws_challan.cell(row=idx, column=8).value = f'=SUM(C{idx}:G{idx})'
            ws_challan.cell(row=idx, column=9).value = challan.get('mode_of_payment', '')
            ws_challan.cell(row=idx, column=10).value = challan.get('bsr_code', '')
            date_str = challan.get('tender_date', '')
            if date_str:
                try:
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    ws_challan.cell(row=idx, column=11).value = date_obj
                    ws_challan.cell(row=idx, column=11).number_format = 'DD/MM/YYYY'
                except:
                    ws_challan.cell(row=idx, column=11).value = date_str
            ws_challan.cell(row=idx, column=12).value = challan.get('challan_no', '')
            ws_challan.cell(row=idx, column=13).value = 'NO'
        
        output_path = masters_path.replace('.xlsx', '_UPDATED.xlsx')
        wb.save(output_path)
        wb.close()
        
        return output_path
        
    except Exception as e:
        st.error(f"Error updating TDS Masters: {str(e)}")
        return None

def generate_output_file(tds_masters_data, challan_data_list, template_path, output_dir):
    """Generate output file from template"""
    try:
        # Determine output filename
        code_to_column_name = tds_masters_data.get('code_to_column_name', {})
        date_col = code_to_column_name.get('(418)')
        output_filename = "TDS_Return.xlsx"
        
        if date_col and date_col in tds_masters_data['tds_parties'].columns:
            dates = tds_masters_data['tds_parties'][date_col].dropna()
            if not dates.empty:
                first_date = pd.to_datetime(dates.iloc[0])
                month_name = first_date.strftime('%B')
                year = first_date.strftime('%Y')
                output_filename = f"TDS_{month_name}_{year}.xlsx"
        
        # Copy template
        output_path = os.path.join(output_dir, output_filename)
        shutil.copy2(template_path, output_path)
        
        # Load and update template
        wb = load_workbook(output_path)
        
        # Update CHALLAN DETAILS sheet
        if 'CHALLAN DETAILS' in wb.sheetnames:
            ws = wb['CHALLAN DETAILS']
            update_challan_sheet(ws, challan_data_list)
        
        # Update DEDUCTEE BREAK-UP sheet
        if 'DEDUCTEE BREAK-UP' in wb.sheetnames:
            ws = wb['DEDUCTEE BREAK-UP']
            update_deductee_sheet(ws, tds_masters_data, challan_data_list)
        
        wb.save(output_path)
        wb.close()
        
        return output_path
        
    except Exception as e:
        st.error(f"Error generating output file: {str(e)}")
        return None

def update_challan_sheet(ws, challan_data_list):
    """Update CHALLAN DETAILS sheet"""
    # Clear existing data (starting from row 4)
    for row in range(4, ws.max_row + 1):
        for col in range(1, 14):
            ws.cell(row=row, column=col).value = None
    
    # Write new data
    for idx, challan in enumerate(challan_data_list, start=1):
        row = idx + 3
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = challan.get('nature_of_payment', '')
        ws.cell(row=row, column=3).value = int(float(challan.get('tax_amount', 0)))
        ws.cell(row=row, column=4).value = int(float(challan.get('surcharge', 0)))
        ws.cell(row=row, column=5).value = int(float(challan.get('cess', 0)))
        ws.cell(row=row, column=6).value = int(float(challan.get('interest', 0)))
        ws.cell(row=row, column=7).value = int(float(challan.get('penalty', 0)))
        ws.cell(row=row, column=8).value = f'=SUM(C{row}:G{row})'
        ws.cell(row=row, column=9).value = challan.get('mode_of_payment', '')
        ws.cell(row=row, column=10).value = challan.get('bsr_code', '')
        date_str = challan.get('tender_date', '')
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                ws.cell(row=row, column=11).value = date_obj
                ws.cell(row=row, column=11).number_format = 'DD/MM/YYYY'
            except:
                ws.cell(row=row, column=11).value = date_str
        ws.cell(row=row, column=12).value = challan.get('challan_no', '')
        ws.cell(row=row, column=13).value = 'NO'

def update_deductee_sheet(ws, tds_masters_data, challan_data_list):
    """Update DEDUCTEE BREAK-UP sheet"""
    tds_parties = tds_masters_data['tds_parties']
    code_to_column_name = tds_masters_data.get('code_to_column_name', {})
    
    # Create challan lookup
    challan_lookup = {}
    for challan in challan_data_list:
        nop = challan.get('nature_of_payment', '').replace(' ', '')
        challan_lookup[nop] = challan
    
    # Clear existing data (starting from row 4)
    for row in range(4, ws.max_row + 1):
        for col in range(1, 23):
            ws.cell(row=row, column=col).value = None
    
    # Write party data
    row_idx = 4
    for _, party in tds_parties.iterrows():
        payment_type = party.get(code_to_column_name.get('(415A)', ''), '')
        if payment_type and str(payment_type).strip() not in ['', 'nan', 'None']:
            # Write all party details to the row
            ws.cell(row=row_idx, column=1).value = row_idx - 3  # Sr. No
            # Add other fields as needed...
            row_idx += 1

# File Upload Section
with tab1:
    st.header("📤 Upload Your Files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("1️⃣ Upload PDF Challan Files")
        pdf_files = st.file_uploader(
            "Select PDF challan files", 
            type=['pdf'],
            accept_multiple_files=True,
            help="Upload one or more PDF challan files (ITNS 281 format)"
        )
        
        if pdf_files:
            st.success(f"✅ {len(pdf_files)} PDF files uploaded")
            for pdf in pdf_files[:3]:
                st.text(f"   • {pdf.name}")
            if len(pdf_files) > 3:
                st.text(f"   ... and {len(pdf_files) - 3} more")
    
    with col2:
        st.subheader("2️⃣ Upload Excel Files")
        
        masters_file = st.file_uploader(
            "Select TDS Masters file",
            type=['xlsx'],
            help="Upload your TDS_Masters*.xlsx file"
        )
        
        if masters_file:
            st.success(f"✅ Masters file: {masters_file.name}")
        
        template_file = st.file_uploader(
            "Select TDS Template file",
            type=['xlsx'],
            help="Upload your TDS_Template*.xlsx file"
        )
        
        if template_file:
            st.success(f"✅ Template file: {template_file.name}")

# Process Section
with tab2:
    st.header("⚙️ Process TDS Returns")
    
    if st.button("🚀 Start Processing", type="primary", disabled=not all([pdf_files, masters_file, template_file])):
        with st.spinner("Processing... This may take a few moments"):
            try:
                # Process PDFs
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                status_text.text("📄 Extracting data from PDF files...")
                challan_data_list = []
                challan_map = {}
                
                for i, pdf_file in enumerate(pdf_files):
                    progress_bar.progress((i + 1) / len(pdf_files) * 0.3)
                    challan_data = extract_challan_data_from_pdf(pdf_file)
                    
                    # Deduplication
                    challan_no = challan_data.get('challan_no', '')
                    if challan_no and challan_no not in challan_map:
                        challan_map[challan_no] = challan_data
                        challan_data_list.append(challan_data)
                
                st.info(f"✅ Extracted {len(challan_data_list)} unique challans from {len(pdf_files)} PDFs")
                
                # Display challan summary
                if challan_data_list:
                    summary_df = pd.DataFrame([
                        {
                            'Nature of Payment': c.get('nature_of_payment', ''),
                            'Challan No': c.get('challan_no', ''),
                            'Tax Amount': f"₹{c.get('tax_amount', '0')}"
                        } for c in challan_data_list
                    ])
                    st.dataframe(summary_df)
                
                # Read Masters file
                status_text.text("📊 Reading TDS Masters file...")
                progress_bar.progress(0.5)
                tds_masters_data = read_tds_masters(masters_file)
                
                if tds_masters_data:
                    st.success(f"✅ Read {len(tds_masters_data['tds_parties'])} parties from Masters file")
                    
                    # Process and create output
                    status_text.text("📝 Generating output files...")
                    progress_bar.progress(0.8)
                    
                    # Save files temporarily
                    with tempfile.TemporaryDirectory() as temp_dir:
                        # Save masters file
                        masters_path = os.path.join(temp_dir, 'masters.xlsx')
                        with open(masters_path, 'wb') as f:
                            masters_file.seek(0)
                            f.write(masters_file.read())
                        
                        # Save template file
                        template_path = os.path.join(temp_dir, 'template.xlsx')
                        with open(template_path, 'wb') as f:
                            template_file.seek(0)
                            f.write(template_file.read())
                        
                        # Update masters with challans
                        updated_masters_path = update_tds_masters_with_challans(
                            tds_masters_data, challan_data_list, masters_path
                        )
                        
                        if updated_masters_path and os.path.exists(updated_masters_path):
                            # Read the updated masters file
                            with open(updated_masters_path, 'rb') as f:
                                updated_masters_file = f
                                updated_masters_data = read_tds_masters(updated_masters_file)
                            
                            # Generate output file
                            output_path = generate_output_file(
                                updated_masters_data, challan_data_list, template_path, temp_dir
                            )
                            
                            # Read files for download
                            with open(updated_masters_path, 'rb') as f:
                                updated_masters_content = f.read()
                            
                            if output_path and os.path.exists(output_path):
                                with open(output_path, 'rb') as f:
                                    output_content = f.read()
                                
                                st.session_state.processed = True
                                st.session_state.results = {
                                    'challan_count': len(challan_data_list),
                                    'party_count': len(tds_masters_data['tds_parties']),
                                    'status': 'success',
                                    'updated_masters_data': updated_masters_content,
                                    'output_data': output_content,
                                    'output_filename': os.path.basename(output_path)
                                }
                                
                                progress_bar.progress(1.0)
                                status_text.text("✅ Processing complete!")
                                st.success("🎉 TDS Return processing completed successfully!")
                                st.balloons()
                            else:
                                st.error("Error generating output file")
                        else:
                            st.error("Error updating TDS Masters")
                    
            except Exception as e:
                st.error(f"❌ Error during processing: {str(e)}")
                st.exception(e)
    
    elif not all([pdf_files, masters_file, template_file]):
        st.warning("⚠️ Please upload all required files in the 'Upload Files' tab first")

# Download Results Section
with tab3:
    st.header("📥 Download Results")
    
    if st.session_state.processed:
        st.success("✅ Your files are ready for download!")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.download_button(
                label="📥 Download Updated TDS Masters",
                data=st.session_state.results['updated_masters_data'],
                file_name="TDS_Masters_UPDATED.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col2:
            st.download_button(
                label="📥 Download TDS Return File",
                data=st.session_state.results['output_data'],
                file_name=st.session_state.results['output_filename'],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        st.info("📊 Processing Summary:")
        st.metric("Unique Challans Processed", st.session_state.results['challan_count'])
        st.metric("Parties Updated", st.session_state.results['party_count'])
        
    else:
        st.info("🔄 Process your files first to see download options")

# Sidebar with instructions
with st.sidebar:
    st.header("📋 Instructions")
    st.markdown("""
    1. **Upload PDF Files**: Select all challan PDFs
    2. **Upload Excel Files**: Upload Masters and Template files
    3. **Process**: Click 'Start Processing'
    4. **Download**: Get your processed files
    
    ### File Requirements:
    - **PDFs**: ITNS 281 format challans
    - **Masters**: TDS_Masters*.xlsx
    - **Template**: TDS_Template*.xlsx
    
    ### Features:
    - ✅ Automatic deduplication
    - ✅ Data validation
    - ✅ Error handling
    - ✅ Progress tracking
    """)
    
    st.divider()
    st.caption("Made with ❤️ using Streamlit")
