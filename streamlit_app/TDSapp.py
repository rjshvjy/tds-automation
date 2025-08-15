# --- Auto-generated Streamlit app (CLEAN) ---
# Source notebook: TDS_Automation_V1.9_without_manual_cleanup.ipynb

import streamlit as st
import os
import io
import re
import time
import glob
from datetime import datetime
import pandas as pd
import numpy as np
import PyPDF2
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

# ====== Original Notebook Functions (embedded) ======
def extract_challan_data_from_pdf(pdf_path):
    """
    Extract challan data from a single PDF file
    Returns a dictionary with all challan details
    """
    challan_data = {}

    try:
        # Open and read the PDF
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)

            # Extract text from first page
            text = ""
            if len(pdf_reader.pages) > 0:
                text = pdf_reader.pages[0].extract_text()

            # Extract all required fields using regex patterns - IMPROVED PATTERNS
            patterns = {
                'tan': r'TAN\s*:\s*([A-Z0-9]+)',
                'nature_of_payment': r'Nature of Payment\s*:\s*(\d+[A-Z])',
                'cin': r'CIN\s*:\s*([A-Z0-9]+)',
                'bsr_code': r'BSR code\s*:\s*([\d]+)',
                'challan_no': r'Challan No\s*:\s*([\d]+)',
                'tender_date': r'Tender Date\s*:\s*(\d{2}/\d{2}/\d{4})',
                'mode_of_payment': r'Mode of Payment\s*:\s*([^\n]+)',
            }

            # Extract each field
            for field, pattern in patterns.items():
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    value = match.group(1).strip()
                    # Convert mode of payment to uppercase
                    if field == 'mode_of_payment':
                        value = value.upper()
                    # Keep BSR code and challan_no as strings to preserve leading zeros
                    elif field in ['bsr_code', 'challan_no']:
                        value = value.zfill(7) if field == 'bsr_code' else value
                    challan_data[field] = value
                else:
                    challan_data[field] = ""

            # SPECIAL HANDLING FOR AMOUNTS - Multiple patterns to try
            # Pattern 1: Try the tax breakup section with flexible whitespace
            tax_patterns = [
                r'A\s+Tax\s+₹\s*([\d,]+)',  # Original pattern
                r'A\s+Tax\s+₹\s*([\d,]+)',  # With regular space
                r'A\s+Tax\s+[₹]\s*([\d,]+)',  # ₹ in brackets
                r'A\s+Tax\s+.\s*([\d,]+)',  # Any character instead of ₹
                r'Tax\s+₹\s*([\d,]+)',  # Simplified pattern
                r'A\s+Tax[^0-9]+([\d,]+)',  # Skip any non-digits after Tax
            ]

            tax_amount = ""
            for pattern in tax_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    tax_amount = match.group(1).strip().replace(',', '')
                    break

            # Fallback: Try the header amount field
            if not tax_amount:
                amount_patterns = [
                    r'Amount \(in Rs\.\)\s*:\s*₹\s*([\d,]+)',
                    r'Amount.*?₹\s*([\d,]+)',
                    r'Amount.*?Rs.*?([\d,]+)',
                ]
                for pattern in amount_patterns:
                    match = re.search(pattern, text, re.IGNORECASE | re.DOTALL)
                    if match:
                        tax_amount = match.group(1).strip().replace(',', '')
                        break

            challan_data['tax_amount'] = tax_amount

            # Extract other amounts with similar flexible patterns
            amount_fields = {
                'surcharge': [
                    r'B\s+Surcharge\s+₹\s*([\d,]+)',
                    r'B\s+Surcharge[^0-9]+([\d,]+)',
                    r'Surcharge\s+₹\s*([\d,]+)'
                ],
                'cess': [
                    r'C\s+Cess\s+₹\s*([\d,]+)',
                    r'C\s+Cess[^0-9]+([\d,]+)',
                    r'Cess\s+₹\s*([\d,]+)'
                ],
                'interest': [
                    r'D\s+Interest\s+₹\s*([\d,]+)',
                    r'D\s+Interest[^0-9]+([\d,]+)',
                    r'Interest\s+₹\s*([\d,]+)'
                ],
                'penalty': [
                    r'E\s+Penalty\s+₹\s*([\d,]+)',
                    r'E\s+Penalty[^0-9]+([\d,]+)',
                    r'Penalty\s+₹\s*([\d,]+)'
                ],
                'fee_234e': [
                    r'F\s+Fee under section 234E\s+₹\s*([\d,]+)',
                    r'Fee under section 234E\s+₹\s*([\d,]+)',
                    r'234E[^0-9]+([\d,]+)'
                ]
            }

            for field, patterns_list in amount_fields.items():
                value = ""
                for pattern in patterns_list:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        value = match.group(1).strip().replace(',', '')
                        break
                challan_data[field] = value if value else "0"

            # Extract total amount
            total_patterns = [
                r'Total \(A\+B\+C\+D\+E\+F\)\s+₹\s*([\d,]+)',
                r'Total.*?₹\s*([\d,]+)',
                r'Total[^0-9]+([\d,]+)'
            ]

            total_amount = ""
            for pattern in total_patterns:
                match = re.search(pattern, text, re.IGNORECASE)
                if match:
                    total_amount = match.group(1).strip().replace(',', '')
                    break

            challan_data['total_amount'] = total_amount

            # Add file name for reference
            challan_data['file_name'] = os.path.basename(pdf_path)

            print(f"✅ Extracted data from: {os.path.basename(pdf_path)}")
            print(f"   Nature of Payment: {challan_data.get('nature_of_payment', 'Not found')}")
            print(f"   Challan No: {challan_data.get('challan_no', 'Not found')}")
            print(f"   Tax Amount: ₹{challan_data.get('tax_amount', 'Not found')}")

            # Debug: Show all amounts if tax amount is found
            if challan_data.get('tax_amount'):
                print(f"   Surcharge: ₹{challan_data.get('surcharge', '0')}")
                print(f"   Cess: ₹{challan_data.get('cess', '0')}")
                print(f"   Total: ₹{challan_data.get('total_amount', 'Not found')}")

    except Exception as e:
        print(f"❌ Error processing {pdf_path}: {str(e)}")
        challan_data['error'] = str(e)

    return challan_data

def extract_all_challans(pdf_folder_path):
    """
    Extract data from all PDF files in a folder and DEDUPLICATE by challan number
    Returns a list of dictionaries, one for each UNIQUE challan
    """
    all_challan_data = []
    challan_map = {}  # To track unique challans by challan number

    # Get all PDF files in the folder
    pdf_files = [f for f in os.listdir(pdf_folder_path) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print(f"❌ No PDF files found in {pdf_folder_path}")
        return all_challan_data

    print(f"\n📁 Found {len(pdf_files)} PDF files to process...")
    print("-" * 50)

    # Process each PDF
    duplicate_count = 0
    for pdf_file in pdf_files:
        pdf_path = os.path.join(pdf_folder_path, pdf_file)
        challan_data = extract_challan_data_from_pdf(pdf_path)

        # Check if this challan number already exists
        challan_no = challan_data.get('challan_no', '')

        if challan_no and challan_no in challan_map:
            # This is a duplicate challan
            duplicate_count += 1
            existing_challan = challan_map[challan_no]

            # Verify the duplicate has same amount (data integrity check)
            if challan_data.get('tax_amount') != existing_challan.get('tax_amount'):
                print(f"⚠️  WARNING: Duplicate challan {challan_no} has different tax amounts!")
                print(f"   File 1: {existing_challan.get('file_name')} - ₹{existing_challan.get('tax_amount')}")
                print(f"   File 2: {challan_data.get('file_name')} - ₹{challan_data.get('tax_amount')}")
        else:
            # This is a new unique challan
            if challan_no:  # Only add if challan number exists
                challan_map[challan_no] = challan_data
                all_challan_data.append(challan_data)
            else:
                print(f"⚠️  Skipping file {pdf_file} - no challan number found")

    print("-" * 50)
    print(f"✅ Total PDF files processed: {len(pdf_files)}")
    print(f"✅ Unique challans found: {len(all_challan_data)}")
    if duplicate_count > 0:
        print(f"ℹ️  Duplicate challans skipped: {duplicate_count}")

    # Create summary by Nature of Payment
    summary = {}
    total_all = 0
    for challan in all_challan_data:
        nop = challan.get('nature_of_payment', 'Unknown')
        if nop:
            if nop not in summary:
                summary[nop] = {'count': 0, 'total_tax': 0}
            summary[nop]['count'] += 1
            try:
                tax_amt = float(challan.get('tax_amount', 0))
                summary[nop]['total_tax'] += tax_amt
                total_all += tax_amt
            except:
                pass

    print("\n📊 Summary by Nature of Payment:")
    for nop, data in sorted(summary.items()):
        print(f"   {nop}: {data['count']} challan(s), Total Tax: ₹{data['total_tax']:,.0f}")
    print(f"   GRAND TOTAL: ₹{total_all:,.0f}")

    return all_challan_data

# Test function to verify extraction is working

def read_tds_masters(file_path):
    """
    Read the TDS Masters Excel file and return data from all sheets
    FIXED:
    1. Removed TDS RATES reading (not used in processing)
    2. Smart row detection - stops at empty data rows (ignores formula-only rows)
    """
    try:
        # Use pandas for TDS CODES only (removed TDS RATES)
        tds_codes = pd.read_excel(file_path, sheet_name='TDS CODES', keep_default_na=False)

        # REMOVED TDS RATES - Not used in processing
        # tds_rates = pd.read_excel(file_path, sheet_name='TDS RATES', keep_default_na=False)

        # Use openpyxl for TDS PARTIES to preserve values
        wb = load_workbook(file_path, data_only=True)
        ws_parties = wb['TDS PARTIES']

        # Find the row with column codes
        code_row = None
        for idx in range(1, 11):  # Check first 10 rows
            row_values = []
            for col in range(1, ws_parties.max_column + 1):
                cell_value = ws_parties.cell(row=idx, column=col).value
                if cell_value:
                    row_values.append(str(cell_value))

            code_patterns = ['(415)', '(427)', '(416)', '(415A)', '-415', '-427', '-416', '-417', '-418', '-419', '-421']
            if any(pattern in val for val in row_values for pattern in code_patterns):
                code_row = idx
                print(f"✅ Found column codes at row {idx}")
                break

        # Read headers (row before codes)
        header_row = code_row - 1 if code_row else 1
        headers = []
        for col in range(1, ws_parties.max_column + 1):
            header_val = ws_parties.cell(row=header_row, column=col).value
            headers.append(header_val if header_val else f"Column_{col}")

        # Get column codes
        code_to_column_name = {}
        column_code_map = {}
        if code_row:
            for col in range(1, ws_parties.max_column + 1):
                code_val = ws_parties.cell(row=code_row, column=col).value
                if code_val:
                    code_str = str(code_val).strip()
                    code_match = None
                    if '(' in code_str and ')' in code_str:
                        code_match = re.search(r'\(([0-9A-Z]+)\)', code_str)
                    elif code_str.startswith('-'):
                        code_match = re.search(r'-([0-9A-Z]+)', code_str)
                    if code_match:
                        extracted_code = code_match.group(1)
                        normalized_code = f'({extracted_code})'
                        column_code_map[normalized_code] = col - 1
                        code_to_column_name[normalized_code] = headers[col - 1]

        # Fallback mappings by column names
        column_name_mappings = {
            '(415)': ['Deductee Code', 'Individual/Company', 'Indiv/Comp', 'Code'],
            '(415A)': ['Section Under Payment Made', 'Type of Payment', 'Nature of Payment', 'Section'],
            '(416)': ['PAN of the Deductee', 'PAN', 'PAN No', 'Deductee PAN'],
            '(417)': ['Name of the Deductee', 'Deductee Name', 'Name', 'Party Name'],
            '(418)': ['Date of Payment/credit', 'Payment Date', 'Date of Payment', 'Credit Date'],
            '(419)': ['Amount Paid /Credited', 'Amount Paid', 'Gross Amount', 'Payment Amount', 'Amount'],
            '(421)': ['TDS', 'Tax Deducted', 'TDS Amount', 'TDS               Rs.', 'TDS Rs.'],
            '(425D)': ['BSR Code', 'BSR', 'Bank BSR Code'],
            '(425E)': ['Challan Serial No', 'Challan No', 'Challan Number'],
            '(425F)': ['Date on which deposited', 'Date Deposited', 'Deposit Date', 'Challan Date'],
            '(427)': ['TDS Deducted Rates %', 'TDS Rate', 'Rate %', 'Deduction Rate', 'Rate']
        }

        for code, possible_names in column_name_mappings.items():
            if code not in code_to_column_name:
                for col_idx, col_name in enumerate(headers):
                    col_name_clean = str(col_name).strip()
                    for possible_name in possible_names:
                        if possible_name.lower() in col_name_clean.lower():
                            code_to_column_name[code] = col_name
                            column_code_map[code] = col_idx
                            print(f"   Found {code} by column name: '{col_name}'")
                            break
                    if code in code_to_column_name:
                        break

        # FIXED: Smart row reading - stop at empty data rows
        data_rows = []
        data_start_row = code_row + 1 if code_row else 2

        # Find critical columns for determining real data
        name_col_idx = None
        pan_col_idx = None

        # Get column indices for Name (417) and PAN (416)
        if '(417)' in code_to_column_name:
            name_col_idx = column_code_map.get('(417)') + 1  # +1 for 1-based Excel columns
        if '(416)' in code_to_column_name:
            pan_col_idx = column_code_map.get('(416)') + 1

        # Count consecutive empty rows
        consecutive_empty = 0
        max_consecutive_empty = 5  # Stop after 5 consecutive empty rows

        for row in range(data_start_row, ws_parties.max_row + 1):
            row_data = []
            has_meaningful_data = False

            # Check if this row has meaningful data
            for col in range(1, ws_parties.max_column + 1):
                cell_value = ws_parties.cell(row=row, column=col).value
                row_data.append(cell_value)

                # Check specifically Name and PAN columns for real data
                if name_col_idx and col == name_col_idx:
                    if cell_value and str(cell_value).strip() and str(cell_value).strip() != '0':
                        has_meaningful_data = True
                elif pan_col_idx and col == pan_col_idx:
                    if cell_value and str(cell_value).strip() and str(cell_value).strip() != '0':
                        has_meaningful_data = True

            # If no Name or PAN columns found, use general check
            if not name_col_idx and not pan_col_idx:
                # Check if at least one cell has non-zero, non-empty value
                for val in row_data:
                    if val is not None and str(val).strip() and str(val).strip() != '0':
                        has_meaningful_data = True
                        break

            # Process the row decision
            if has_meaningful_data:
                data_rows.append(row_data)
                consecutive_empty = 0  # Reset counter
            else:
                consecutive_empty += 1
                # Stop if we've seen enough consecutive empty rows
                if consecutive_empty >= max_consecutive_empty:
                    print(f"   Stopped reading at row {row} (found {consecutive_empty} consecutive empty rows)")
                    break

        print(f"   Read {len(data_rows)} rows with actual data (ignoring formula-only rows)")

        tds_parties = pd.DataFrame(data_rows, columns=headers)

        # Convert numeric columns with precise rounding - EXCLUDE (427) TO PRESERVE DECIMAL RATES
        numeric_codes = ['(419)', '(421)']  # Removed '(427)' to avoid quantizing rates to integers
        for code in numeric_codes:
            col_name = code_to_column_name.get(code)
            if col_name and col_name in tds_parties.columns:
                tds_parties[col_name] = pd.to_numeric(
                    tds_parties[col_name].astype(str).str.replace(',', '').str.replace('₹', '').str.strip(),
                    errors='coerce'
                ).apply(lambda x: Decimal(str(x)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if pd.notna(x) else x)
                print(f"   ✅ Converted '{col_name}' to numeric with ROUND_HALF_UP")

        # Convert date columns
        date_codes = ['(418)', '(425F)']
        for code in date_codes:
            col_name = code_to_column_name.get(code)
            if col_name and col_name in tds_parties.columns:
                tds_parties[col_name] = pd.to_datetime(tds_parties[col_name], errors='coerce', dayfirst=True)
                print(f"   ✅ Converted '{col_name}' to datetime")

        # Validate PANs (only for rows with actual data)
        pan_col = code_to_column_name.get('(416)')
        if pan_col and pan_col in tds_parties.columns:
            invalid_pan_count = 0
            for idx, pan in tds_parties[pan_col].items():
                # Only validate if PAN exists and is not empty/zero
                if pd.notna(pan) and str(pan).strip() and str(pan).strip() != '0':
                    if not re.match(r'^[A-Z]{5}[0-9]{4}[A-Z]$', str(pan)):
                        invalid_pan_count += 1
                        if invalid_pan_count <= 5:  # Only show first 5 warnings
                            print(f"⚠️ Invalid PAN format at row {idx + data_start_row}: {pan}")
            if invalid_pan_count > 5:
                print(f"   ... and {invalid_pan_count - 5} more invalid PANs")

        challan_details = pd.read_excel(file_path, sheet_name='Challan Details', header=1, keep_default_na=False)
        wb.close()

        print(f"\n✅ Successfully read TDS Masters file")
        print(f"   TDS PARTIES: {len(tds_parties)} rows (actual data only)")
        print(f"   Challan Details: {len(challan_details)} rows")
        print(f"   TDS CODES: {len(tds_codes)} entries")
        # REMOVED: print(f"   TDS RATES: {len(tds_rates)} rates")
        print(f"   Column codes mapped: {len(code_to_column_name)}")

        tds_col = code_to_column_name.get('(421)', None)
        if tds_col and tds_col in tds_parties.columns:
            print(f"\n📊 Sample TDS amounts from first 5 rows:")
            for idx in range(min(5, len(tds_parties))):
                payment_type = tds_parties.iloc[idx].get(code_to_column_name.get('(415A)', ''), '')
                tds_amount = tds_parties.iloc[idx].get(tds_col, 0)
                name = tds_parties.iloc[idx].get(code_to_column_name.get('(417)', ''), '')
                print(f"   Row {idx}: Name={name}, Payment={payment_type}, TDS={tds_amount}")

        print("\n📊 Column Code Mapping Found:")
        for code, col_name in sorted(code_to_column_name.items())[:15]:
            print(f"   Code {code} → Column: '{col_name}'")

        return {
            'tds_codes': tds_codes,
            'tds_parties': tds_parties,
            'challan_details': challan_details,
            # REMOVED: 'tds_rates': tds_rates,
            'file_path': file_path,
            'column_code_map': column_code_map,
            'code_to_column_name': code_to_column_name,
            'code_row': code_row
        }

    except Exception as e:
        print(f"❌ Error reading TDS Masters: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def update_tds_masters_with_challans(tds_masters_data, challan_data_list):
    """
    Update TDS Masters with challan information
    FIXED: Uses data_only=True to preserve static TDS values, writes BSR/challan as strings
    """
    try:
        # Load workbook with data_only=True to preserve static values
        wb = load_workbook(tds_masters_data['file_path'], data_only=True)
        ws_parties = wb['TDS PARTIES']
        ws_challan = wb['Challan Details']

        code_to_column_name = tds_masters_data.get('code_to_column_name', {})
        code_row = tds_masters_data.get('code_row', 1)

        # Find columns by codes
        col_425E = col_425F = col_415A = None
        for col_idx in range(1, ws_parties.max_column + 1):
            cell_value = str(ws_parties.cell(row=code_row, column=col_idx).value)
            if '425E' in cell_value:
                col_425E = col_idx
                print(f"Found (425E) at column {col_idx}")
            elif '425F' in cell_value:
                col_425F = col_idx
                print(f"Found (425F) at column {col_idx}")
            elif '415A' in cell_value:
                col_415A = col_idx
                print(f"Found (415A) (Type of Payment) at column {col_idx}")

        # Create mapping of nature of payment to challan data
        challan_map = {}
        for challan in challan_data_list:
            nop = challan.get('nature_of_payment', '')
            if nop:
                nop_clean = nop.replace(' ', '')
                challan_map[nop_clean] = challan

        print(f"\n📝 Updating TDS PARTIES sheet...")
        print(f"   Challan Serial No (425E) → Column {col_425E}")
        print(f"   Date deposited (425F) → Column {col_425F}")
        updates_made = 0

        # Update TDS PARTIES
        data_start_row = code_row + 1
        for row_idx in range(data_start_row, ws_parties.max_row + 1):
            payment_type = ws_parties.cell(row=row_idx, column=col_415A).value if col_415A else None
            if payment_type and str(payment_type).strip() not in ['', 'nan', 'None']:
                payment_type_clean = str(payment_type).replace(' ', '').strip()
                if payment_type_clean in challan_map:
                    challan = challan_map[payment_type_clean]
                    if col_425E:
                        ws_parties.cell(row=row_idx, column=col_425E).value = challan.get('challan_no', '')
                    if col_425F:
                        date_str = challan.get('tender_date', '')
                        if date_str:
                            try:
                                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                                ws_parties.cell(row=row_idx, column=col_425F).value = date_obj
                                ws_parties.cell(row=row_idx, column=col_425F).number_format = 'DD/MM/YYYY'
                            except:
                                ws_parties.cell(row=row_idx, column=col_425F).value = date_str
                    updates_made += 1

        print(f"✅ Updated {updates_made} rows in TDS PARTIES")

        # Update Challan Details
        print("\n📝 Updating Challan Details sheet...")
        for row in ws_challan.iter_rows(min_row=3, max_row=ws_challan.max_row):
            for cell in row:
                cell.value = None

        for idx, challan in enumerate(challan_data_list, start=3):
            tax_amt = Decimal(challan.get('tax_amount', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('tax_amount', '') else 0
            surcharge = Decimal(challan.get('surcharge', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('surcharge', '') else 0
            cess = Decimal(challan.get('cess', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('cess', '') else 0
            interest = Decimal(challan.get('interest', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('interest', '') else 0
            penalty = Decimal(challan.get('penalty', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('penalty', '') else 0

            ws_challan.cell(row=idx, column=1).value = idx - 2
            ws_challan.cell(row=idx, column=2).value = challan.get('nature_of_payment', '')
            ws_challan.cell(row=idx, column=3).value = int(tax_amt)
            ws_challan.cell(row=idx, column=4).value = int(surcharge)
            ws_challan.cell(row=idx, column=5).value = int(cess)
            ws_challan.cell(row=idx, column=6).value = int(interest)
            ws_challan.cell(row=idx, column=7).value = int(penalty)
            ws_challan.cell(row=idx, column=8).value = f'=SUM(C{idx}:G{idx})'
            ws_challan.cell(row=idx, column=9).value = challan.get('mode_of_payment', '')
            ws_challan.cell(row=idx, column=10).value = challan.get('bsr_code', '')
            date_str = challan.get('tender_date', '')
            if date_str:
                try:
                    date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                    ws_challan.cell(row=idx, column=11).value = date_obj
                    ws_challan.cell(row=idx, column=11).number_format = 'DD/MM/YYYY'
                except:
                    ws_challan.cell(row=idx, column=11).value = date_str
            ws_challan.cell(row=idx, column=12).value = challan.get('challan_no', '')
            ws_challan.cell(row=idx, column=13).value = 'NO'

        print(f"✅ Added {len(challan_data_list)} challans to Challan Details")

        output_file = tds_masters_data['file_path'].replace('.xlsx', '_UPDATED.xlsx')
        wb.save(output_file)
        wb.close()

        print(f"\n✅ Saved updated TDS Masters to: {output_file}")
        return read_tds_masters(output_file)

    except Exception as e:
        print(f"❌ Error updating TDS Masters: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def validate_tds_totals(tds_masters_data, challan_data_list):
    """
    Validate that party-wise TDS totals match challan amounts
    Uses column codes to identify the correct columns
    """
    print("\n🔍 Validating TDS totals...")

    try:
        tds_parties = tds_masters_data['tds_parties']
        code_to_column_name = tds_masters_data.get('code_to_column_name', {})
        payment_col = code_to_column_name.get('(415A)', None)
        tds_col = code_to_column_name.get('(421)', None)

        print(f"\n📊 Debug - Column mappings:")
        print(f"   Payment Type column (415A): {payment_col}")
        print(f"   TDS Amount column (421): {tds_col}")

        if not payment_col or not tds_col:
            print("⚠️ Missing required columns")
            return False

        print(f"\n📊 Debug - Sample data (first 5 valid rows):")
        sample_count = 0
        for idx, row in tds_parties.iterrows():
            if sample_count >= 5:
                break
            payment = row.get(payment_col, '') if payment_col else ''
            tds_amount = row.get(tds_col, 0) if tds_col else 0
            if payment and str(payment) not in ['nan', 'NaT', '']:
                print(f"   Row {idx}: Payment={payment}, TDS={tds_amount}")
                sample_count += 1

        party_totals = {}
        for _, row in tds_parties.iterrows():
            payment_type = str(row[payment_col] if payment_col in row else '').strip()
            if payment_type and payment_type != 'nan' and payment_type != 'NaT':
                tds_amount = 0
                if tds_col and tds_col in row:
                    try:
                        val = row[tds_col]
                        if pd.notna(val):
                            tds_amount = Decimal(str(val)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP)
                    except Exception as e:
                        print(f"   Warning: Could not convert TDS value '{val}' for payment type {payment_type}: {e}")
                        tds_amount = 0
                payment_type_clean = payment_type.replace(' ', '')
                if payment_type_clean not in party_totals:
                    party_totals[payment_type_clean] = 0
                party_totals[payment_type_clean] += tds_amount

        challan_totals = {}
        for challan in challan_data_list:
            nop = challan.get('nature_of_payment', '').replace(' ', '')
            tax_amount = Decimal(challan.get('tax_amount', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('tax_amount', '') else 0
            if nop:
                challan_totals[nop] = tax_amount

        print(f"\n📊 Debug - Totals found:")
        print(f"   Party totals: {party_totals}")
        print(f"   Challan totals: {challan_totals}")

        validation_passed = True
        print("\n📊 Validation Results:")
        print("-" * 60)
        print(f"{'Nature of Payment':<20} {'Party Total':<15} {'Challan Total':<15} {'Status':<10}")
        print("-" * 60)

        for nop in sorted(set(list(party_totals.keys()) + list(challan_totals.keys()))):
            party_total = party_totals.get(nop, 0)
            challan_total = challan_totals.get(nop, 0)
            difference = abs(party_total - challan_total)
            status = "✅ PASS" if difference <= 1 else "❌ FAIL"
            if difference > 1:
                validation_passed = False
            print(f"{nop:<20} ₹{party_total:<14,.0f} ₹{challan_total:<14,.0f} {status}")

        print("-" * 60)
        print("\n✅ All validations passed!" if validation_passed else "\n❌ Validation failed! Please check the discrepancies above.")

        return validation_passed

    except Exception as e:
        print(f"❌ Error during validation: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

print("✅ Excel handling functions loaded - COMPLETE FIX!")
print("   ✓ REMOVED TDS RATES reading (not used in processing)")
print("   ✓ Smart row detection - stops at empty data rows")
print("   ✓ Ignores formula-only rows with 0 or empty values")
print("   ✓ Uses Name (417) and PAN (416) to detect real data")
print("   ✓ Stops after 5 consecutive empty rows")
print("   ✓ Only validates PANs for rows with actual data")
print("   ✓ All other functionality preserved")

# Cell 4: Main processing function that coordinates all steps
# This cell contains the main process_tds_returns function

import math
from decimal import Decimal, ROUND_HALF_UP

def find_totals_row(ws, start_row=4, end_row=None):
    """
    Find the row containing totals (usually has 'TOTAL' text or SUM formulas)
    """
    if end_row is None:
        end_row = ws.max_row

    for row in range(start_row, end_row + 1):
        # Check first few columns for 'TOTAL' text
        for col in range(1, 4):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and 'total' in str(cell_value).lower():
                return row

        # Also check if row has SUM formulas
        cell_formula = ws.cell(row=row, column=3).value
        if cell_formula and isinstance(cell_formula, str) and cell_formula.startswith('=SUM'):
            return row

    return None

def update_challan_details(ws, challan_data_list):
    """
    Helper function to update CHALLAN DETAILS sheet
    """
    print("\n📝 Updating CHALLAN DETAILS...")
    print(f"   Processing {len(challan_data_list)} unique challans...")

    # Find totals row
    totals_row = None
    for row in range(4, ws.max_row + 1):
        if str(ws.cell(row=row, column=2).value).lower() == 'total':
            totals_row = row
            break

    if not totals_row:
        totals_row = 8  # Default to row 8 if not found (based on template)
    print(f"   Found TOTAL row at row {totals_row}")

    data_start = 4
    available_slots = totals_row - data_start

    print(f"   Template has space for {available_slots} data rows")
    print(f"   Need {len(challan_data_list)} rows for unique challans")

    # Insert rows if needed
    if len(challan_data_list) > available_slots:
        rows_to_insert = len(challan_data_list) - available_slots
        ws.insert_rows(totals_row, amount=rows_to_insert)
        totals_row += rows_to_insert
        print(f"   Inserted {rows_to_insert} additional rows. New TOTAL row: {totals_row}")

    # Clear existing data
    for row in ws.iter_rows(min_row=4, max_row=totals_row-1):
        for cell in row:
            cell.value = None

    # Write challans
    row_idx = 4
    for idx, challan in enumerate(challan_data_list, start=1):
        tax_amt = Decimal(challan.get('tax_amount', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('tax_amount', '') else 0
        surcharge = Decimal(challan.get('surcharge', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('surcharge', '') else 0
        cess = Decimal(challan.get('cess', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('cess', '') else 0
        interest = Decimal(challan.get('interest', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('interest', '') else 0
        penalty = Decimal(challan.get('penalty', 0)).quantize(Decimal('1.'), rounding=ROUND_HALF_UP) if challan.get('penalty', '') else 0

        ws.cell(row=row_idx, column=1).value = idx
        ws.cell(row=row_idx, column=2).value = challan.get('nature_of_payment', '')
        ws.cell(row=row_idx, column=3).value = int(tax_amt)
        ws.cell(row=row_idx, column=4).value = int(surcharge)
        ws.cell(row=row_idx, column=5).value = int(cess)
        ws.cell(row=row_idx, column=6).value = int(interest)
        ws.cell(row=row_idx, column=7).value = int(penalty)
        ws.cell(row=row_idx, column=8).value = f'=SUM(C{row_idx}:G{row_idx})'
        ws.cell(row=row_idx, column=9).value = challan.get('mode_of_payment', '')
        ws.cell(row=row_idx, column=10).value = challan.get('bsr_code', '')
        date_str = challan.get('tender_date', '')
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                ws.cell(row=row_idx, column=11).value = date_obj
                ws.cell(row=row_idx, column=11).number_format = 'DD/MM/YYYY'
            except:
                ws.cell(row=row_idx, column=11).value = date_str
        ws.cell(row=row_idx, column=12).value = challan.get('challan_no', '')
        ws.cell(row=row_idx, column=13).value = 'NO'

        row_idx += 1

    # Setup TOTAL row
    ws.cell(row=totals_row, column=2).value = 'TOTAL'
    ws.cell(row=totals_row, column=3).value = f'=SUM(C{data_start}:C{totals_row-1})'
    ws.cell(row=totals_row, column=4).value = f'=SUM(D{data_start}:D{totals_row-1})'
    ws.cell(row=totals_row, column=5).value = f'=SUM(E{data_start}:E{totals_row-1})'
    ws.cell(row=totals_row, column=6).value = f'=SUM(F{data_start}:F{totals_row-1})'
    ws.cell(row=totals_row, column=7).value = f'=SUM(G{data_start}:G{totals_row-1})'
    ws.cell(row=totals_row, column=8).value = f'=SUM(H{data_start}:H{totals_row-1})'

    # Clear any rows after totals to prevent overflow
    if ws.max_row > totals_row:
        delete_amount = ws.max_row - totals_row
        ws.delete_rows(totals_row + 1, delete_amount)
        print(f"   Cleared {delete_amount} overflow rows after totals in CHALLAN DETAILS. Final max_row: {ws.max_row}")

    print(f"✅ Successfully updated CHALLAN DETAILS with {len(challan_data_list)} unique challans")
    print(f"   TOTAL row is at row {totals_row}")

def update_challan_details_proper(ws, challan_data_list):
    """
    Update the CHALLAN DETAILS sheet with proper TOTAL row handling
    FIXED: Preserves TOTAL row when clearing data and writes BSR/Challan as strings
    """
    # CRITICAL: Headers are in rows 1-3, data starts at row 4
    DATA_START_ROW = 4

    # Step 1: Find the current TOTAL row
    total_row = find_totals_row(ws, DATA_START_ROW)

    if total_row:
        print(f"   Found TOTAL row at row {total_row}")
        current_data_rows = total_row - DATA_START_ROW
        print(f"   Template has space for {current_data_rows} data rows")
    else:
        print("   No TOTAL row found - will add at end")
        current_data_rows = 0
        total_row = None

    # Step 2: Calculate how many rows we need (based on unique challans)
    needed_rows = len(challan_data_list)
    print(f"   Need {needed_rows} rows for unique challans")

    # Step 3: Adjust rows if needed
    if total_row and needed_rows > current_data_rows:
        # We need to insert rows BEFORE the total row
        rows_to_insert = needed_rows - current_data_rows
        print(f"   Inserting {rows_to_insert} rows before TOTAL row")
        ws.insert_rows(total_row, rows_to_insert)
        # Update total row position
        total_row = total_row + rows_to_insert

    elif total_row and needed_rows < current_data_rows:
        # We need to delete extra rows
        rows_to_delete = current_data_rows - needed_rows
        print(f"   Deleting {rows_to_delete} extra rows")
        # Delete rows just before the total row
        for _ in range(rows_to_delete):
            ws.delete_rows(total_row - 1)
            total_row -= 1

    # Step 4: Clear ONLY data cells (not the total row)
    print(f"   Clearing data from row {DATA_START_ROW} to row {DATA_START_ROW + needed_rows - 1}")
    end_clear_row = DATA_START_ROW + needed_rows
    if total_row:
        # Make sure we don't clear the total row
        end_clear_row = min(end_clear_row, total_row)

    for row in range(DATA_START_ROW, end_clear_row):
        for col in range(1, 14):  # Columns A to M
            ws.cell(row=row, column=col).value = None

    # Step 5: Write the data (only unique challans)
    current_row = DATA_START_ROW
    for idx, challan in enumerate(challan_data_list, start=1):
        # Round amounts UP to nearest rupee
        tax_amt = math.ceil(float(challan.get('tax_amount', 0))) if challan.get('tax_amount', '') else 0
        surcharge = math.ceil(float(challan.get('surcharge', 0))) if challan.get('surcharge', '') else 0
        cess = math.ceil(float(challan.get('cess', 0))) if challan.get('cess', '') else 0
        interest = math.ceil(float(challan.get('interest', 0))) if challan.get('interest', '') else 0
        penalty = math.ceil(float(challan.get('penalty', 0))) if challan.get('penalty', '') else 0

        # Write data to proper columns
        ws.cell(row=current_row, column=1).value = idx  # (401) SR NO

        # (402) SECTION CODE - Format with space
        section_code = challan.get('nature_of_payment', '')
        if len(section_code) >= 3 and ' ' not in section_code:
            section_code = section_code[:2] + ' ' + section_code[2:]
        ws.cell(row=current_row, column=2).value = section_code

        ws.cell(row=current_row, column=3).value = tax_amt  # (403) TDS Rs.
        ws.cell(row=current_row, column=4).value = surcharge  # (404) SURCHARGE Rs.
        ws.cell(row=current_row, column=5).value = cess  # (405) EDUCATION CESS Rs.
        ws.cell(row=current_row, column=6).value = interest  # (406) INTEREST Rs.
        ws.cell(row=current_row, column=7).value = penalty  # (407) OTHERS Rs.

        # (408) TOTAL TAX DEPOSITED - Formula for column H
        ws.cell(row=current_row, column=8).value = f'=SUM(C{current_row}:G{current_row})'

        # (409) CHEQUE/DD NO - Mode of payment already uppercase from PDF extraction
        ws.cell(row=current_row, column=9).value = challan.get('mode_of_payment', '')

        # (410) BSR CODE - as string to preserve leading zeros
        ws.cell(row=current_row, column=10).value = challan.get('bsr_code', '')

        # (411) DATE ON WHICH TAX DEPOSITED - parse and format date
        date_str = challan.get('tender_date', '')
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                ws.cell(row=current_row, column=11).value = date_obj
                ws.cell(row=current_row, column=11).number_format = 'DD/MM/YYYY'
            except:
                ws.cell(row=current_row, column=11).value = date_str

        # (412) CHALLAN SERIAL NO - as string to preserve leading zeros
        ws.cell(row=current_row, column=12).value = challan.get('challan_no', '')

        ws.cell(row=current_row, column=13).value = 'NO'  # (413) WHETHER TDS DEPOSITED BY BOOK ENTRY

        current_row += 1

    # Step 6: Create or Update TOTAL row
    if not total_row:
        # Create new total row if it doesn't exist
        total_row = DATA_START_ROW + needed_rows

    print(f"   Setting up TOTAL row at row {total_row}")

    # Ensure TOTAL text is in column A
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    # Update formulas to cover the correct range
    last_data_row = DATA_START_ROW + needed_rows - 1

    # Update formulas for columns C through H (only where totals exist in template)
    for col in [3, 4, 5, 6, 7, 8]:  # Columns C, D, E, F, G, H
        ws.cell(row=total_row, column=col).value = f'=SUM({get_column_letter(col)}{DATA_START_ROW}:{get_column_letter(col)}{last_data_row})'
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    print(f"✅ Successfully updated CHALLAN DETAILS with {needed_rows} unique challans")
    print(f"   TOTAL row is at row {total_row}")

def update_deductee_breakup(ws, tds_masters_data, challan_data_list):
    """
    Helper function to update DEDUCTEE BREAK-UP sheet
    FIXED: Properly detect totals row by checking for formulas or standard template position
    """
    code_to_column_name = tds_masters_data.get('code_to_column_name', {})
    tds_parties = tds_masters_data['tds_parties']
    code_row = tds_masters_data.get('code_row', 1)

    col_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = str(ws.cell(row=2, column=col_idx).value)
        if '414' in cell_value:
            col_map['sr_no'] = col_idx
        elif '415' in cell_value and '415A' not in cell_value:
            col_map['deductee_code'] = col_idx
        elif '415A' in cell_value:
            col_map['payment_type'] = col_idx
        elif '416' in cell_value:
            col_map['pan'] = col_idx
        elif '417' in cell_value:
            col_map['name'] = col_idx
        elif '418' in cell_value:
            col_map['date_payment'] = col_idx
        elif '419' in cell_value:
            col_map['amount_paid'] = col_idx
        elif '420' in cell_value:
            col_map['book_entry'] = col_idx
        elif '421' in cell_value:
            col_map['tds'] = col_idx
        elif '422' in cell_value:
            col_map['surcharge'] = col_idx
        elif '423' in cell_value:
            col_map['cess'] = col_idx
        elif '424' in cell_value:
            col_map['total_deducted'] = col_idx
        elif '425' in cell_value and '425A' not in cell_value and '425B' not in cell_value and '425C' not in cell_value and '425D' not in cell_value and '425E' not in cell_value and '425F' not in cell_value:
            col_map['total_deposited'] = col_idx
        elif '425A' in cell_value:
            col_map['interest'] = col_idx
        elif '425B' in cell_value:
            col_map['others'] = col_idx
        elif '425C' in cell_value:
            col_map['total'] = col_idx
        elif '425D' in cell_value:
            col_map['bsr_code'] = col_idx
        elif '425E' in cell_value:
            col_map['challan_no'] = col_idx
        elif '425F' in cell_value:
            col_map['date_deposited'] = col_idx
        elif '426' in cell_value:
            col_map['date_deduction'] = col_idx
        elif '427' in cell_value:
            col_map['rate'] = col_idx
        elif '428' in cell_value:
            col_map['reason'] = col_idx

    print("\n📝 Updating DEDUCTEE BREAK-UP...")
    print("   Column Mappings Found:")
    for key, col in col_map.items():
        print(f"   - {key}: {ws.cell(row=1, column=col).value}")

    # Create challan map
    challan_map = {challan.get('nature_of_payment', '').replace(' ', ''): challan for challan in challan_data_list}

    # FIXED: Find totals row by checking for formulas or standard position
    totals_row = None
    data_start = 4

    # Method 1: Check for SUM formulas in column 7 (Amount Paid column)
    for row in range(data_start, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=7).value
        if cell_value is not None:
            if isinstance(cell_value, str) and '=SUM' in cell_value.upper():
                totals_row = row
                print(f"   Found totals row at row {totals_row} (has SUM formula)")
                break

    # Method 2: Check row 55 specifically (standard template position)
    if not totals_row and ws.max_row >= 55:
        # Check if row 55 has formulas or all zeros
        row_55_has_formulas = False
        row_55_all_zeros = True
        for col in range(7, 17):  # Check numeric columns G to P
            cell_val = ws.cell(row=55, column=col).value
            if cell_val is not None:
                if isinstance(cell_val, str) and '=' in cell_val:
                    row_55_has_formulas = True
                elif cell_val != 0:
                    row_55_all_zeros = False

        if row_55_has_formulas or row_55_all_zeros:
            totals_row = 55
            print(f"   Found totals row at row {totals_row} (standard template position)")

    # Method 3: Look for row with all zeros after row 20
    if not totals_row:
        for row in range(20, ws.max_row + 1):
            all_zeros = True
            has_values = False
            for col in range(7, 17):  # Check numeric columns
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    has_values = True
                    if val != 0 and not (isinstance(val, str) and '=' in val):
                        all_zeros = False
                        break

            if has_values and all_zeros:
                # Verify next row is empty
                next_row_empty = all(ws.cell(row=row+1, column=c).value is None for c in range(1, 10))
                if next_row_empty or row == ws.max_row:
                    totals_row = row
                    print(f"   Found totals row at row {totals_row} (detected by pattern)")
                    break

    # Fallback
    if not totals_row:
        totals_row = 55  # Default to row 55 based on template
        print(f"   Using default totals row position: {totals_row}")

    # Count valid parties
    valid_parties = []
    payment_col = code_to_column_name.get('(415A)', '')
    for _, party in tds_parties.iterrows():
        payment_type = party.get(payment_col, '') if payment_col else ''
        if payment_type and str(payment_type).strip() not in ['', 'nan', 'None']:
            valid_parties.append(party)

    print(f"   Found {len(valid_parties)} valid parties to process")

    # Clear existing data rows (but not totals row)
    for row in ws.iter_rows(min_row=data_start, max_row=totals_row-1):
        for cell in row:
            cell.value = None

    # Dynamic row management - insert or delete rows as needed
    available_slots = totals_row - data_start
    if len(valid_parties) > available_slots:
        rows_to_insert = len(valid_parties) - available_slots
        ws.insert_rows(totals_row, amount=rows_to_insert)
        totals_row += rows_to_insert
        print(f"   Inserted {rows_to_insert} additional rows. New totals row: {totals_row}")
    elif len(valid_parties) < available_slots - 5:  # Only delete if significantly fewer rows needed
        rows_to_delete = available_slots - len(valid_parties) - 5  # Keep 5 buffer rows
        if rows_to_delete > 0:
            ws.delete_rows(totals_row - rows_to_delete, rows_to_delete)
            totals_row -= rows_to_delete
            print(f"   Deleted {rows_to_delete} excess rows. New totals row: {totals_row}")

    # Write parties
    row_idx = data_start
    party_count = 0
    for _, party in tds_parties.iterrows():
        payment_type = party.get(payment_col, '') if payment_col else ''
        if payment_type and str(payment_type).strip() not in ['', 'nan', 'None']:
            payment_type_clean = str(payment_type).replace(' ', '').strip()
            challan = challan_map.get(payment_type_clean, {})

            # Write each field
            if 'sr_no' in col_map:
                ws.cell(row=row_idx, column=col_map['sr_no']).value = party_count + 1

            if 'deductee_code' in col_map:
                deductee_code = party.get(code_to_column_name.get('(415)', ''), '') if code_to_column_name.get('(415)') else ''
                ws.cell(row=row_idx, column=col_map['deductee_code']).value = deductee_code

            if 'payment_type' in col_map:
                ws.cell(row=row_idx, column=col_map['payment_type']).value = payment_type

            if 'pan' in col_map:
                pan = party.get(code_to_column_name.get('(416)', ''), '') if code_to_column_name.get('(416)') else ''
                ws.cell(row=row_idx, column=col_map['pan']).value = pan

            if 'name' in col_map:
                name = party.get(code_to_column_name.get('(417)', ''), '') if code_to_column_name.get('(417)') else ''
                ws.cell(row=row_idx, column=col_map['name']).value = name

            if 'date_payment' in col_map:
                date_payment = party.get(code_to_column_name.get('(418)', ''), '') if code_to_column_name.get('(418)') else ''
                if isinstance(date_payment, pd.Timestamp):
                    date_payment = date_payment.to_pydatetime()
                ws.cell(row=row_idx, column=col_map['date_payment']).value = date_payment
                ws.cell(row=row_idx, column=col_map['date_payment']).number_format = 'DD/MM/YYYY'

            if 'amount_paid' in col_map:
                amount_paid = party.get(code_to_column_name.get('(419)', ''), 0) if code_to_column_name.get('(419)') else 0
                if isinstance(amount_paid, Decimal):
                    amount_paid = int(amount_paid)
                ws.cell(row=row_idx, column=col_map['amount_paid']).value = amount_paid

            if 'book_entry' in col_map:
                ws.cell(row=row_idx, column=col_map['book_entry']).value = ''

            if 'tds' in col_map:
                tds_amount = party.get(code_to_column_name.get('(421)', ''), 0) if code_to_column_name.get('(421)') else 0
                if isinstance(tds_amount, Decimal):
                    tds_amount = int(tds_amount)
                ws.cell(row=row_idx, column=col_map['tds']).value = tds_amount

            if 'surcharge' in col_map:
                ws.cell(row=row_idx, column=col_map['surcharge']).value = 0

            if 'cess' in col_map:
                ws.cell(row=row_idx, column=col_map['cess']).value = 0

            if 'total_deducted' in col_map:
                ws.cell(row=row_idx, column=col_map['total_deducted']).value = f'=SUM(I{row_idx}:K{row_idx})'

            if 'total_deposited' in col_map:
                ws.cell(row=row_idx, column=col_map['total_deposited']).value = ws.cell(row=row_idx, column=col_map['total_deducted']).value

            if 'interest' in col_map:
                ws.cell(row=row_idx, column=col_map['interest']).value = 0

            if 'others' in col_map:
                ws.cell(row=row_idx, column=col_map['others']).value = 0

            if 'total' in col_map:
                ws.cell(row=row_idx, column=col_map['total']).value = f'=SUM(M{row_idx}:O{row_idx})'

            if 'bsr_code' in col_map:
                ws.cell(row=row_idx, column=col_map['bsr_code']).value = challan.get('bsr_code', '')

            if 'challan_no' in col_map:
                ws.cell(row=row_idx, column=col_map['challan_no']).value = challan.get('challan_no', '')

            if 'date_deposited' in col_map:
                date_str = challan.get('tender_date', '')
                if date_str:
                    try:
                        date_obj = datetime.strptime(date_str, '%d/%m/%Y')
                        ws.cell(row=row_idx, column=col_map['date_deposited']).value = date_obj
                        ws.cell(row=row_idx, column=col_map['date_deposited']).number_format = 'DD/MM/YYYY'
                    except:
                        ws.cell(row=row_idx, column=col_map['date_deposited']).value = date_str

            if 'date_deduction' in col_map:
                date_deduction = party.get(code_to_column_name.get('(418)', ''), '') if code_to_column_name.get('(418)') else ''
                if isinstance(date_deduction, pd.Timestamp):
                    date_deduction = date_deduction.to_pydatetime()
                ws.cell(row=row_idx, column=col_map['date_deduction']).value = date_deduction
                ws.cell(row=row_idx, column=col_map['date_deduction']).number_format = 'DD/MM/YYYY'

            if 'rate' in col_map:
                rate = party.get(code_to_column_name.get('(427)', ''), 0) if code_to_column_name.get('(427)') else 0
                rate_percent = float(rate) * 100
                ws.cell(row=row_idx, column=col_map['rate']).value = f"{rate_percent:.2f}%"

            if 'reason' in col_map:
                ws.cell(row=row_idx, column=col_map['reason']).value = 'N.A'

            row_idx += 1
            party_count += 1

    # DO NOT add "TOTAL" text - template doesn't have it
    # Just update the formulas in the totals row
    print(f"   Updating formulas in totals row at row {totals_row}")

    # Sum formulas for numeric columns
    if 'amount_paid' in col_map:
        ws.cell(row=totals_row, column=col_map['amount_paid']).value = f'=SUM(G{data_start}:G{row_idx-1})'

    if 'tds' in col_map:
        ws.cell(row=totals_row, column=col_map['tds']).value = f'=SUM(I{data_start}:I{row_idx-1})'

    if 'surcharge' in col_map:
        ws.cell(row=totals_row, column=col_map['surcharge']).value = f'=SUM(J{data_start}:J{row_idx-1})'

    if 'cess' in col_map:
        ws.cell(row=totals_row, column=col_map['cess']).value = f'=SUM(K{data_start}:K{row_idx-1})'

    if 'total_deducted' in col_map:
        ws.cell(row=totals_row, column=col_map['total_deducted']).value = f'=SUM(L{data_start}:L{row_idx-1})'

    if 'total_deposited' in col_map:
        ws.cell(row=totals_row, column=col_map['total_deposited']).value = f'=SUM(M{data_start}:M{row_idx-1})'

    if 'interest' in col_map:
        ws.cell(row=totals_row, column=col_map['interest']).value = f'=SUM(N{data_start}:N{row_idx-1})'

    if 'others' in col_map:
        ws.cell(row=totals_row, column=col_map['others']).value = f'=SUM(O{data_start}:O{row_idx-1})'

    if 'total' in col_map:
        ws.cell(row=totals_row, column=col_map['total']).value = f'=SUM(P{data_start}:P{row_idx-1})'

    # Clear any rows after totals to prevent overflow
    if ws.max_row > totals_row:
        delete_amount = ws.max_row - totals_row
        ws.delete_rows(totals_row + 1, delete_amount)
        print(f"   Cleared {delete_amount} overflow rows after totals. Final max_row: {ws.max_row}")

    print(f"✅ Updated {party_count} parties in DEDUCTEE BREAK-UP")
    print(f"   Totals row is at row {totals_row}")

def update_deductee_breakup_sheet_dynamic(ws, tds_masters_data, challan_data_list):
    """
    Update the DEDUCTEE BREAK-UP sheet with dynamic row management
    FIXED: Properly handles rate formatting and preserves leading zeros
    """
    # Get TDS parties data and column mappings
    tds_parties_df = tds_masters_data['tds_parties']
    code_to_column_name = tds_masters_data.get('code_to_column_name', {})

    # Find columns by their codes
    col_name = code_to_column_name.get('(417)', None)  # Name
    col_pan = code_to_column_name.get('(416)', None)  # PAN
    col_type_payment = code_to_column_name.get('(415A)', None)  # Type of Payment
    col_code_415 = code_to_column_name.get('(415)', None)  # Individual/Company Code
    col_date_payment = code_to_column_name.get('(418)', None)  # Date of Payment
    col_amount = code_to_column_name.get('(419)', None)  # Amount Paid
    col_tds = code_to_column_name.get('(421)', None)  # TDS Amount
    col_tds_rate = code_to_column_name.get('(427)', None)  # TDS Deduction Rates
    col_bsr = code_to_column_name.get('(425D)', None)  # BSR Code
    col_challan_no = code_to_column_name.get('(425E)', None)  # Challan No
    col_date_deposited = code_to_column_name.get('(425F)', None)  # Date deposited

    # Debug: Print column mappings
    print(f"   Column Mappings Found:")
    print(f"   - Name (417): {col_name}")
    print(f"   - PAN (416): {col_pan}")
    print(f"   - Type of Payment (415A): {col_type_payment}")
    print(f"   - Code (415): {col_code_415}")
    print(f"   - Date of Payment (418): {col_date_payment}")
    print(f"   - Amount Paid (419): {col_amount}")
    print(f"   - TDS (421): {col_tds}")
    print(f"   - TDS Rate (427): {col_tds_rate}")
    print(f"   - BSR Code (425D): {col_bsr}")
    print(f"   - Challan No (425E): {col_challan_no}")
    print(f"   - Date Deposited (425F): {col_date_deposited}")

    # Find where data should start (after headers and column codes)
    DATA_START_ROW = 4

    # Count actual data rows needed (skip empty payment types)
    if col_type_payment:
        # Filter for valid payment types (should be like '94A', '94C', etc.)
        valid_parties = tds_parties_df[
            (tds_parties_df[col_type_payment].notna()) &
            (tds_parties_df[col_type_payment].astype(str).str.strip() != '') &
            (tds_parties_df[col_type_payment].astype(str) != 'nan')
        ]
    else:
        valid_parties = tds_parties_df

    needed_rows = len(valid_parties)
    print(f"   Found {needed_rows} valid parties to process")

    # Find the totals row
    total_row = None
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and 'total' in str(cell_value).lower():
            total_row = row
            break
        # Also check for SUM formulas in column G (Amount column)
        cell_g = ws.cell(row=row, column=7).value
        if cell_g and isinstance(cell_g, str) and '=SUM' in str(cell_g):
            total_row = row
            break

    if total_row:
        print(f"   Found totals row at row {total_row}")

        # Calculate how many data rows we currently have
        current_data_rows = total_row - DATA_START_ROW

        # Adjust rows to exactly match needed rows
        if needed_rows > current_data_rows:
            # Insert rows
            rows_to_insert = needed_rows - current_data_rows
            print(f"   Inserting {rows_to_insert} rows before TOTAL row")
            ws.insert_rows(total_row, rows_to_insert)
            total_row = total_row + rows_to_insert
        elif needed_rows < current_data_rows:
            # Delete excess rows
            rows_to_delete = current_data_rows - needed_rows
            print(f"   Deleting {rows_to_delete} excess rows")
            for _ in range(rows_to_delete):
                ws.delete_rows(total_row - 1)
                total_row -= 1

    # Clear existing data (but NOT the totals row)
    end_clear_row = DATA_START_ROW + needed_rows
    if total_row:
        end_clear_row = min(end_clear_row, total_row)

    for row in range(DATA_START_ROW, end_clear_row):
        for col in range(1, 23):  # Clear all columns A to V
            ws.cell(row=row, column=col).value = None

    # Create challan lookup for BSR codes
    challan_lookup = {}
    for challan in challan_data_list:
        nop = challan.get('nature_of_payment', '').replace(' ', '')
        challan_lookup[nop] = {
            'bsr_code': challan.get('bsr_code', ''),
            'challan_no': challan.get('challan_no', ''),
            'date_deposited': challan.get('tender_date', '')
        }

    # Add party data
    current_row = DATA_START_ROW
    sr_no = 1

    for _, party in valid_parties.iterrows():
        # (414) Sr.No
        ws.cell(row=current_row, column=1).value = sr_no

        # (415) Deductee Code - from column (415)
        deductee_code = ''
        if col_code_415:
            deductee_code = str(party.get(col_code_415, '')).strip()
            # Ensure deductee code is properly formatted (should be 01 or 02 with leading zero)
            if deductee_code and deductee_code.isdigit():
                deductee_code = deductee_code.zfill(2)  # Add leading zero if needed
        else:
            # Default based on PAN 4th character
            if col_pan:
                pan_value = str(party.get(col_pan, '')).strip()
                if len(pan_value) >= 4:
                    fourth_char = pan_value[3].upper()
                    deductee_code = '01' if fourth_char == 'P' else '02'
                else:
                    deductee_code = '01'
            else:
                deductee_code = '01'
        ws.cell(row=current_row, column=2).value = deductee_code

      # (415A) Section Under Payment Made - Format with space
        payment_type = str(party.get(col_type_payment, '')) if col_type_payment else ''
        payment_type_clean = payment_type.replace(' ', '')  # For lookup
        if payment_type and payment_type not in ['nan', 'None', '']:
            # Format payment type with space between number and letter
            payment_type = payment_type.strip()
            if len(payment_type) >= 3 and payment_type[:2].isdigit() and payment_type[2].isalpha():
                # Add space between number and letter (e.g., "94A" -> "94 A")
                formatted_payment = payment_type[:2] + ' ' + payment_type[2:]
            else:
                formatted_payment = payment_type
            ws.cell(row=current_row, column=3).value = formatted_payment

        # (416) PAN of Deductee
        pan_value = party.get(col_pan, '') if col_pan else ''
        ws.cell(row=current_row, column=4).value = pan_value

        # (417) Name of Deductee
        name_value = party.get(col_name, '') if col_name else ''
        ws.cell(row=current_row, column=5).value = name_value

        # (418) Date of Payment/credit
        if col_date_payment:
            date_val = party.get(col_date_payment)
            if pd.notna(date_val):
                ws.cell(row=current_row, column=6).value = date_val
                ws.cell(row=current_row, column=6).number_format = 'DD/MM/YYYY'

        # (419) Amount Paid/Credited Rs. - Round UP
        amount = 0
        if col_amount:
            amount_val = party.get(col_amount)
            if pd.notna(amount_val):
                try:
                    amount = math.ceil(float(amount_val))
                except:
                    amount = 0
        ws.cell(row=current_row, column=7).value = amount

        # (420) Paid by Book Entry or otherwise - Left blank as requested
        # ws.cell(row=current_row, column=8).value = ''

        # (421) TDS Rs. - Round UP
        tds = 0
        if col_tds:
            tds_val = party.get(col_tds)
            if pd.notna(tds_val):
                try:
                    tds = math.ceil(float(tds_val))
                except:
                    tds = 0
        ws.cell(row=current_row, column=9).value = tds

        # (422) Surcharge Rs.
        ws.cell(row=current_row, column=10).value = 0

        # (423) Educational Cess Rs.
        ws.cell(row=current_row, column=11).value = 0

        # (424) Total tax deducted - Formula
        ws.cell(row=current_row, column=12).value = f'=I{current_row}+J{current_row}+K{current_row}'

        # (425) Total tax deposited Rs. - Same as total tax deducted
        ws.cell(row=current_row, column=13).value = f'=L{current_row}'

        # (425A) Interest
        ws.cell(row=current_row, column=14).value = 0

        # (425B) Others
        ws.cell(row=current_row, column=15).value = 0

        # (425C) Total (425+Interest+Others) - Formula
        ws.cell(row=current_row, column=16).value = f'=M{current_row}+N{current_row}+O{current_row}'

        # (425D, 425E, 425F) - Get from updated TDS Masters or challan lookup
        challan_info = challan_lookup.get(payment_type_clean, {})

        # First check if data exists in party row (from updated TDS Masters)
        bsr_value = party.get(col_bsr, '') if col_bsr else ''
        challan_no_value = party.get(col_challan_no, '') if col_challan_no else ''
        date_dep_value = party.get(col_date_deposited, '') if col_date_deposited else ''

        # If not in party data, get from challan lookup
        if not bsr_value:
            bsr_value = challan_info.get('bsr_code', '')
        if not challan_no_value:
            challan_no_value = challan_info.get('challan_no', '')
        if not date_dep_value:
            date_dep_value = challan_info.get('date_deposited', '')

        # Write BSR code as string to preserve leading zeros
        ws.cell(row=current_row, column=17).value = str(bsr_value)

        # Write challan no as string to preserve leading zeros
        ws.cell(row=current_row, column=18).value = str(challan_no_value)

        # Write date deposited
        if date_dep_value:
            if isinstance(date_dep_value, str):
                try:
                    date_obj = datetime.strptime(date_dep_value, '%d/%m/%Y')
                    ws.cell(row=current_row, column=19).value = date_obj
                    ws.cell(row=current_row, column=19).number_format = 'DD/MM/YYYY'
                except:
                    ws.cell(row=current_row, column=19).value = date_dep_value
            else:
                ws.cell(row=current_row, column=19).value = date_dep_value
                ws.cell(row=current_row, column=19).number_format = 'DD/MM/YYYY'

        # (426) Date of deduction - Same as payment date
        if col_date_payment:
            date_val = party.get(col_date_payment)
            if pd.notna(date_val):
                ws.cell(row=current_row, column=20).value = date_val
                ws.cell(row=current_row, column=20).number_format = 'DD/MM/YYYY'

        # (427) Rate at which deducted - FIXED FORMATTING
        if col_tds_rate:
            tds_rate_value = party.get(col_tds_rate)
            if pd.notna(tds_rate_value):
                # Convert to string and clean
                rate_str = str(tds_rate_value).strip()
                # Remove % if present
                rate_str = rate_str.replace('%', '')
                # Try to convert to float to validate it's a number
                try:
                    rate_float = float(rate_str)
                    # If rate is decimal (like 0.1), multiply by 100 to get percentage
                    if rate_float < 1:
                        rate_float = rate_float * 100
                    # Format consistently
                    ws.cell(row=current_row, column=21).value = f'{rate_float:.2f}%'
                except:
                    # If conversion fails, use as is with % appended
                    ws.cell(row=current_row, column=21).value = f'{rate_str}%'
            else:
                # If no rate found, calculate from TDS/Amount
                if amount > 0 and tds > 0:
                    rate = (tds / amount) * 100
                    ws.cell(row=current_row, column=21).value = f'{rate:.2f}%'
                else:
                    ws.cell(row=current_row, column=21).value = '0%'
        else:
            # Fallback: Calculate rate if TDS rate column not found
            if amount > 0 and tds > 0:
                rate = (tds / amount) * 100
                ws.cell(row=current_row, column=21).value = f'{rate:.2f}%'
            else:
                ws.cell(row=current_row, column=21).value = '0%'

        # (428) Reason for non-deduction/lower deduction
        ws.cell(row=current_row, column=22).value = 'N.A'

        current_row += 1
        sr_no += 1

    # Create or Update totals row
    if not total_row:
        total_row = DATA_START_ROW + needed_rows

    print(f"   Setting up totals row at row {total_row}")

    last_data_row = DATA_START_ROW + needed_rows - 1

    # Ensure TOTAL text is in column A
    ws.cell(row=total_row, column=1).value = "TOTAL"
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    # Update sum formulas for numeric columns (only where totals exist in template)
    total_columns = [
        (7, 'G'),   # Amount
        (9, 'I'),   # TDS
        (10, 'J'),  # Surcharge
        (11, 'K'),  # Cess
        (12, 'L'),  # Total tax deducted
        (13, 'M'),  # Total tax deposited
        (14, 'N'),  # Interest
        (15, 'O'),  # Others
        (16, 'P')   # Total
    ]

    for col, col_letter in total_columns:
        ws.cell(row=total_row, column=col).value = f'=SUM({col_letter}{DATA_START_ROW}:{col_letter}{last_data_row})'
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    print(f"✅ Updated {sr_no - 1} parties in DEDUCTEE BREAK-UP")
    print(f"   Totals row is at row {total_row}")

print("✅ Output file generation functions loaded - COMPLETE FIX!")
print("   ✓ TOTAL rows are preserved in both sheets")
print("   ✓ All columns properly mapped from TDS Masters")
print("   ✓ BSR Code, Challan No written as strings to preserve leading zeros")
print("   ✓ Rate formatting fixed - decimals converted to percentages (0.1 → 10%)")
print("   ✓ Dynamic row management preserves TOTAL formulas")
print("   ✓ Column H (Paid by Book Entry) left blank as requested")
print("   ✓ Proper date formatting and string conversion where appropriate")

# Cell 6: Testing individual components and troubleshooting - FIXED VERSION
# Use these functions to test each step separately if you encounter issues

def get_output_filename_from_masters(tds_masters_data):
    """
    Extract month and year from the first payment date in TDS Masters
    to generate output filename as TDS_Month_Year.xlsx
    """
    try:
        code_to_column_name = tds_masters_data.get('code_to_column_name', {})
        date_col = code_to_column_name.get('(418)')
        if date_col and date_col in tds_masters_data['tds_parties'].columns:
            dates = tds_masters_data['tds_parties'][date_col].dropna()
            if not dates.empty:
                first_date = pd.to_datetime(dates.iloc[0])
                month_name = first_date.strftime('%B')
                year = first_date.strftime('%Y')
                return f"TDS_{month_name}_{year}.xlsx"
        current_date = datetime.now()
        return f"TDS_{current_date.strftime('%B')}_{current_date.strftime('%Y')}.xlsx"
    except:
        current_date = datetime.now()
        return f"TDS_{current_date.strftime('%B')}_{current_date.strftime('%Y')}.xlsx"

def generate_output_file(tds_masters_data, challan_data_list, template_path, output_path=None):
    """
    Generate output file from TDS Masters data and challan information
    Handles dynamic rows and preserves formulas
    """
    try:
        # Load the template
        wb = load_workbook(template_path)

        # Get the sheets
        ws_deductor = wb['DEDUCTOR DETAILS']
        ws_challan = wb['CHALLAN DETAILS']
        ws_deductee = wb['DEDUCTEE BREAK-UP']

        print("✅ Loaded template")

        # Update CHALLAN DETAILS sheet with deduplicated challans
        print("\n📝 Updating CHALLAN DETAILS...")
        print(f"   Processing {len(challan_data_list)} unique challans...")
        update_challan_details_proper(ws_challan, challan_data_list)

        # Update DEDUCTEE BREAK-UP sheet
        print("\n📝 Updating DEDUCTEE BREAK-UP...")
        update_deductee_breakup_sheet_dynamic(ws_deductee, tds_masters_data, challan_data_list)

        # Save the file
        wb.save(output_path)
        wb.close()

        print(f"\n✅ Generated output file: {output_path}")
        return output_path

    except Exception as e:
        print(f"❌ Error generating output file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# ====== Helpers ======
def save_uploaded_files(pdf_files, masters_file, template_file, workdir):
    pdf_dir = os.path.join(workdir, 'pdfs')
    os.makedirs(pdf_dir, exist_ok=True)
    saved_pdfs = []
    for up in pdf_files or []:
        path = os.path.join(pdf_dir, up.name)
        with open(path, 'wb') as f:
            f.write(up.getbuffer())
        saved_pdfs.append(path)
    masters_path = None
    if masters_file is not None:
        masters_path = os.path.join(workdir, masters_file.name)
        with open(masters_path, 'wb') as f:
            f.write(masters_file.getbuffer())
    template_path = None
    if template_file is not None:
        template_path = os.path.join(workdir, template_file.name)
        with open(template_path, 'wb') as f:
            f.write(template_file.getbuffer())
    return saved_pdfs, masters_path, template_path

def streamlit_process(pdf_paths, masters_path, template_path, outdir):
    challans = extract_all_challans(os.path.dirname(pdf_paths[0])) if pdf_paths else []
    masters = read_tds_masters(masters_path)
    updated = update_tds_masters_with_challans(masters, challans)
    out_name = get_output_filename_from_masters(updated)
    out_path = os.path.join(outdir, out_name)
    gen_info = generate_output_file(updated, challans, template_path, output_path=out_path)
    if isinstance(gen_info, dict):
        gen_info['output_path'] = gen_info.get('output_path', out_path)
        return gen_info
    return {'output_path': out_path}

# ====== Streamlit UI ======
st.set_page_config(page_title='TDS Automation', page_icon='🧾', layout='wide')
st.title('🧾 TDS Automation — Streamlit App')

with st.sidebar:
    st.header('How it works')
    st.markdown(
        """
        1) Upload **PDF challans** (one or many)  
        2) Upload **TDS Masters** Excel  
        3) Upload **Template** Excel  
        4) Click **Process** and download the result
        """
    )
    
    st.divider()
    st.header('📥 Download Templates')
    st.markdown(
        """
        **Required files from Google Drive:**
        
        📄 [**Help File (PDF)**](https://drive.google.com/uc?export=download&id=1ns1cfmFlLJbkxfYD3ScKCbxYZbQj4gNG)  
        *Instructions and guidelines*
        
        📊 [**TDS_Masters Template**](https://docs.google.com/spreadsheets/d/1WogLXgbftVWblZTnYhwhKjxPcHjnY77G/export?format=xlsx)  
        *Fill this file and upload*
        
        📋 [**TDS_Template**](https://docs.google.com/spreadsheets/d/1LQ_3tR72c1xGry64_lQeaC4EiUeqSZ2x/export?format=xlsx)  
        *Fill DEDUCTOR DETAILS and upload*
        """
    )
    
    st.caption('If download button is missing, see the Debug box or check the Results folder.')
st.subheader('1) Upload files')
pdf_files = st.file_uploader('PDF Challans (multiple allowed)', type=['pdf'], accept_multiple_files=True)
masters_file = st.file_uploader('TDS Masters (Excel)', type=['xlsx', 'xls'])
template_file = st.file_uploader('Template (Excel)', type=['xlsx', 'xls'])

with st.expander('🔧 Debug (optional)'):
    st.write('PDFs:', [f.name for f in (pdf_files or [])])
    st.write('Masters:', getattr(masters_file, 'name', None))
    st.write('Template:', getattr(template_file, 'name', None))

st.divider()
st.subheader('2) Process')
go = st.button('🚀 Process')

if go:
    if not pdf_files or masters_file is None or template_file is None:
        st.warning('Please upload PDFs, a Masters Excel, and a Template Excel.')
        st.stop()

    with st.status('Processing…', expanded=True) as status:
        workdir = 'workdir'
        outdir = os.path.join(workdir, 'output')
        os.makedirs(outdir, exist_ok=True)
        st.write('Saving uploaded files…')
        saved_pdfs, masters_path, template_path = save_uploaded_files(pdf_files, masters_file, template_file, workdir)
        status.update(label='Running pipeline…')
        output_path = None
        try:
            result = streamlit_process(saved_pdfs, masters_path, template_path, outdir)
            if isinstance(result, dict) and 'output_path' in result:
                output_path = result['output_path']
            if not output_path:
                candidates = sorted(glob.glob(os.path.join(outdir, '*.xls*')), key=os.path.getmtime, reverse=True)
                if candidates:
                    output_path = candidates[0]
            status.update(label='Done!', state='complete', expanded=False)
        except Exception as e:
            status.update(label='Failed', state='error')
            st.exception(e)
            st.stop()

    if not output_path or not os.path.exists(output_path):
        st.error('Processing finished but no Excel file was found.')
    else:
        # Find the updated masters file
        updated_masters_path = None
        for file in glob.glob(os.path.join(workdir, '*_UPDATED.xlsx')):
            updated_masters_path = file
            break
        
        st.success('✅ Processing Complete! Download your files below:')
        
        # Create two columns for download buttons
        col1, col2 = st.columns(2)
        
        # Download button for Updated Masters
        if updated_masters_path and os.path.exists(updated_masters_path):
            with col1:
                with open(updated_masters_path, 'rb') as f:
                    st.download_button(
                        '📊 Download Updated Masters', 
                        f, 
                        file_name=os.path.basename(updated_masters_path),
                        key='masters_download'
                    )
        
        # Download button for Output file
        with col2:
            with open(output_path, 'rb') as f:
                st.download_button(
                    '📋 Download TDS Return File', 
                    f, 
                    file_name=os.path.basename(output_path),
                    key='output_download'
                )

st.caption('Tip: Keep file names and sheet names as in your existing workflow.')
